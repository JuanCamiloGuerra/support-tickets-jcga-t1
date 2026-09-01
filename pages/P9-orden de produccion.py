import math
import re
import unicodedata
from io import BytesIO
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)


BASE_DIR = Path(__file__).resolve().parents[1]
VENTAS_PATH = BASE_DIR / "tablas" / "ventas_df.csv"
INVENTARIO_PATH = BASE_DIR / "tablas" / "inventario.csv"

TARGET_YEAR = 2027
TARGET_MONTHS = [1, 2]
TARGET_PERIOD_LABEL = "enero-febrero 2027"

TARGET_SCHOOLS = [
    "liceo pupo jimenez",
    "liceo universitario",
    "nuestra señora del carmen",
    "gimnasio plaza feliz",
    "gimnasio vallegrande",
    "colsafa",
    "conalco",
    "la inmaculada",
    "Antonio Narino",
    "hogar",
    "otros",
    "Pantalones Azules",
    "Bermudas Azules",
]

EXCLUDED_PRODUCT_WORDS = ["medias"]
BOGOTA_TZ = ZoneInfo("America/Bogota")


st.set_page_config(
    page_title="Orden de produccion",
    page_icon="📋",
    layout="wide",
)

st.title("📋 Orden de produccion enero-febrero 2027")

components.html(
    """
    <script>
        setTimeout(function() {
            window.parent.location.reload();
        }, 300000);
    </script>
    """,
    height=0,
)


def normalize_text(value):
    text = str(value or "").strip().lower()
    text = text.replace("�", "ñ")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_col(value):
    return normalize_text(value).replace(" ", "_")


def clean_number(value):
    if pd.isna(value):
        return 0.0

    text = str(value).strip()
    if text == "":
        return 0.0

    text = text.replace("$", "").replace(" ", "")

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return 0.0


def sort_size(value):
    text = str(value).strip()
    lower = text.lower()

    if lower == "unica":
        return (2, 0, lower)

    if lower.isdigit():
        return (0, int(lower), lower)

    alpha_order = {
        "s": 100,
        "m": 101,
        "l": 102,
        "xl": 103,
        "xxl": 104,
    }
    return (1, alpha_order.get(lower, 999), lower)


def get_column(df, candidates):
    normalized = {
        normalize_col(column): column
        for column in df.columns
    }

    for candidate in candidates:
        key = normalize_col(candidate)
        if key in normalized:
            return normalized[key]

    raise KeyError(f"No se encontro la columna: {candidates}")


def get_optional_column(df, candidates):
    try:
        return get_column(df, candidates)
    except KeyError:
        return None


def round_up_to_5(value):
    if value <= 0:
        return 0
    return int(math.ceil(value / 5) * 5)


def now_bogota():
    return datetime.now(BOGOTA_TZ)


def file_updated_at_bogota(path):
    return datetime.fromtimestamp(path.stat().st_mtime, BOGOTA_TZ)


def format_bogota(dt_value):
    return dt_value.astimezone(BOGOTA_TZ).strftime("%Y-%m-%d %H:%M:%S")


def load_data():
    ventas = pd.read_csv(
        VENTAS_PATH,
        sep=";",
        dtype=str,
        keep_default_na=False,
    )
    inventario = pd.read_csv(
        INVENTARIO_PATH,
        sep=";",
        dtype=str,
        keep_default_na=False,
    )

    ventas.columns = ventas.columns.str.strip()
    inventario.columns = inventario.columns.str.strip()

    sales_cols = {
        "school": get_column(ventas, ["colegio"]),
        "product": get_column(ventas, ["articulo", "producto"]),
        "size": get_column(ventas, ["talla"]),
        "qty": get_column(ventas, ["cantidad"]),
        "month": get_column(ventas, ["mes"]),
        "year": get_column(ventas, ["año", "ano", "a�o"]),
    }
    sales_id_col = get_optional_column(
        ventas,
        [
            "ID unico de articulo",
            "ID unico de artículo",
            "ID unico de artñculo",
            "ID_BUSQUEDA",
        ],
    )
    inv_cols = {
        "school": get_column(inventario, ["colegio"]),
        "product": get_column(inventario, ["producto"]),
        "size": get_column(inventario, ["talla"]),
        "stock": get_column(inventario, ["inventario"]),
    }
    inventory_id_col = get_optional_column(inventario, ["ID_BUSQUEDA", "ID unico de articulo"])
    production_col = get_optional_column(inventario, ["produccion", "producción"])

    ventas = ventas.rename(columns={value: key for key, value in sales_cols.items()})
    inventario = inventario.rename(columns={value: key for key, value in inv_cols.items()})
    if sales_id_col:
        ventas = ventas.rename(columns={sales_id_col: "item_id"})
    else:
        ventas["item_id"] = ""
    if inventory_id_col:
        inventario = inventario.rename(columns={inventory_id_col: "item_id"})
    else:
        inventario["item_id"] = ""
    if production_col:
        inventario = inventario.rename(columns={production_col: "production_family"})
    else:
        inventario["production_family"] = ""

    ventas["school_norm"] = ventas["school"].map(normalize_text)
    ventas["product_norm"] = ventas["product"].map(normalize_text)
    ventas["size_norm"] = ventas["size"].replace("", "UNICA").astype(str).str.strip().str.upper()
    ventas.loc[ventas["size_norm"].isin(["", "NAN", "NONE"]), "size_norm"] = "UNICA"
    ventas["item_id_norm"] = ventas["item_id"].map(normalize_text)
    ventas["qty"] = ventas["qty"].map(clean_number)
    ventas["month"] = ventas["month"].map(clean_number).astype(int)
    ventas["year"] = ventas["year"].map(clean_number).astype(int)

    inventario["school_norm"] = inventario["school"].map(normalize_text)
    inventario["product_norm"] = inventario["product"].map(normalize_text)
    inventario["size_norm"] = inventario["size"].replace("", "UNICA").astype(str).str.strip().str.upper()
    inventario.loc[inventario["size_norm"].isin(["", "NAN", "NONE"]), "size_norm"] = "UNICA"
    inventario["item_id_norm"] = inventario["item_id"].map(normalize_text)
    inventario["familia_produccion"] = inventario["production_family"].astype(str).str.strip()
    inventario["familia_produccion_norm"] = inventario["familia_produccion"].map(normalize_text)
    inventario["stock"] = inventario["stock"].map(clean_number).clip(lower=0)

    aliases = {
        "inmaculada": "la inmaculada",
        "gimnsio plaza feliz": "gimnasio plaza feliz",
    }
    ventas["school_norm"] = ventas["school_norm"].replace(aliases)
    inventario["school_norm"] = inventario["school_norm"].replace(aliases)
    inventario = inventario[inventario["familia_produccion_norm"].ne("")].copy()

    id_family_map = (
        inventario[inventario["item_id_norm"].ne("")]
        .drop_duplicates("item_id_norm")
        .set_index("item_id_norm")["familia_produccion"]
        .to_dict()
    )
    key_family_map = (
        inventario
        .drop_duplicates(["school_norm", "product_norm", "size_norm"])
        .set_index(["school_norm", "product_norm", "size_norm"])["familia_produccion"]
        .to_dict()
    )

    family_by_key = pd.Series(
        ventas.set_index(["school_norm", "product_norm", "size_norm"]).index.map(key_family_map),
        index=ventas.index,
    )
    ventas["familia_produccion"] = ventas["item_id_norm"].map(id_family_map)
    ventas["familia_produccion"] = ventas["familia_produccion"].fillna(family_by_key)
    ventas["familia_produccion_norm"] = ventas["familia_produccion"].map(normalize_text)
    ventas = ventas[ventas["familia_produccion_norm"].ne("")].copy()

    school_norms = {
        normalize_text(school)
        for school in TARGET_SCHOOLS
    }
    excluded_pattern = "|".join(EXCLUDED_PRODUCT_WORDS)

    ventas = ventas[
        ventas["school_norm"].isin(school_norms)
        & ~ventas["product_norm"].str.contains(excluded_pattern, na=False)
    ].copy()
    inventario = inventario[
        inventario["school_norm"].isin(school_norms)
        & ~inventario["product_norm"].str.contains(excluded_pattern, na=False)
    ].copy()

    return ventas, inventario


def month_forecast(monthly, key, month):
    subset = monthly[
        (monthly["month"] == month)
        & (monthly["school_norm"] == key["school_norm"])
        & (monthly["product_norm"] == key["product_norm"])
        & (monthly["size_norm"] == key["size_norm"])
        & (monthly["familia_produccion"] == key["familia_produccion"])
    ]["qty"]

    if subset.empty:
        avg_qty = 0.0
        p75_qty = 0.0
        std_qty = 0.0
        recent_qty = 0.0
    else:
        avg_qty = float(subset.mean())
        p75_qty = float(subset.quantile(0.75))
        std_qty = float(subset.std(ddof=0))
        recent = monthly[
            (monthly["month"] == month)
            & (monthly["year"] == TARGET_YEAR - 1)
            & (monthly["school_norm"] == key["school_norm"])
            & (monthly["product_norm"] == key["product_norm"])
            & (monthly["size_norm"] == key["size_norm"])
            & (monthly["familia_produccion"] == key["familia_produccion"])
        ]["qty"]
        recent_qty = float(recent.iloc[0]) if not recent.empty else 0.0

    weighted_recent = (recent_qty * 0.6) + (avg_qty * 0.4)
    forecast = math.ceil(max(avg_qty, p75_qty, weighted_recent))
    safety = math.ceil(max(forecast * 0.2, std_qty)) if forecast else 0
    return forecast, safety


def build_order():
    ventas, inventario = load_data()

    history = ventas[
        (ventas["year"] < TARGET_YEAR)
        & (ventas["month"].isin(TARGET_MONTHS))
    ]
    monthly = (
        history
        .groupby(
            ["month", "year", "school_norm", "product_norm", "size_norm", "familia_produccion"],
            dropna=False,
        )["qty"]
        .sum()
        .reset_index()
    )

    current_year = (
        ventas[
            (ventas["year"] == TARGET_YEAR)
            & (ventas["month"].isin(TARGET_MONTHS))
        ]
        .groupby(
            ["month", "school_norm", "product_norm", "size_norm", "familia_produccion"],
            dropna=False,
        )["qty"]
        .sum()
        .reset_index(name="current_qty")
    )

    inventory_group = (
        inventario
        .groupby(
            ["school_norm", "product_norm", "size_norm", "familia_produccion"],
            dropna=False,
        )["stock"]
        .sum()
        .reset_index()
    )

    keys = pd.concat(
        [
            monthly[["school_norm", "product_norm", "size_norm", "familia_produccion"]],
            current_year[["school_norm", "product_norm", "size_norm", "familia_produccion"]],
            inventory_group[["school_norm", "product_norm", "size_norm", "familia_produccion"]],
        ],
        ignore_index=True,
    ).drop_duplicates()

    rows = []

    for _, key in keys.iterrows():
        forecast_total = 0
        safety_total = 0

        for month in TARGET_MONTHS:
            forecast, safety = month_forecast(monthly, key, month)
            current_match = current_year[
                (current_year["month"] == month)
                & (current_year["school_norm"] == key["school_norm"])
                & (current_year["product_norm"] == key["product_norm"])
                & (current_year["size_norm"] == key["size_norm"])
                & (current_year["familia_produccion"] == key["familia_produccion"])
            ]["current_qty"]
            current_qty = float(current_match.iloc[0]) if not current_match.empty else 0.0

            forecast_total += max(forecast, math.ceil(current_qty))
            safety_total += safety

        stock_match = inventory_group[
            (inventory_group["school_norm"] == key["school_norm"])
            & (inventory_group["product_norm"] == key["product_norm"])
            & (inventory_group["size_norm"] == key["size_norm"])
            & (inventory_group["familia_produccion"] == key["familia_produccion"])
        ]["stock"]
        stock = float(stock_match.iloc[0]) if not stock_match.empty else 0.0

        raw_order = max(0, forecast_total + safety_total - stock)
        production = round_up_to_5(raw_order)

        if production > 0:
            rows.append(
                {
                    "colegio": key["school_norm"],
                    "articulo": key["product_norm"],
                    "familia_produccion": key["familia_produccion"],
                    "colegio_prenda": f"{key['school_norm']} | {key['product_norm']}",
                    "talla": str(key["size_norm"]),
                    "inventario": int(round(stock)),
                    "pronostico_cierre": int(forecast_total),
                    "stock_minimo": int(safety_total),
                    "orden_sin_redondeo": int(math.ceil(raw_order)),
                    "orden_produccion": production,
                }
            )

    detail = pd.DataFrame(rows)
    if detail.empty:
        return detail, pd.DataFrame()

    detail = detail.sort_values(
        ["familia_produccion", "colegio", "articulo", "talla"],
        key=lambda col: col.astype(str),
    )

    return detail, build_pivot(detail)


def build_pivot(detail):
    if detail.empty:
        return pd.DataFrame()

    pivot = detail.pivot_table(
        index=["colegio", "colegio_prenda"],
        columns="talla",
        values="orden_produccion",
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.sort_index(level=[0, 1])
    pivot = pivot[[column for column in sorted(pivot.columns, key=sort_size)]]
    pivot.insert(0, "colegio y prenda", [idx[1] for idx in pivot.index])
    pivot = pivot.reset_index(drop=True)

    return pivot


def color_order_cells(value):
    if not isinstance(value, (int, float)):
        return ""

    if value > 15:
        return "background-color:#f8d7da;color:#8a1f2d;font-weight:bold;"

    if value > 10:
        return "background-color:#fff3cd;color:#664d03;font-weight:bold;"

    return ""


def add_total_row(df, label_column):
    if df.empty:
        return df

    numeric_columns = [
        column
        for column in df.columns
        if column != label_column
    ]
    total_row = {column: int(df[column].sum()) for column in numeric_columns}
    total_row[label_column] = "TOTAL POR TALLA"
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def style_total_row(row):
    if str(row.iloc[0]).strip().upper() != "TOTAL POR TALLA":
        return [""] * len(row)

    return [
        "background-color:#243b53;color:white;font-weight:bold;"
        for _ in row
    ]


def pdf_page_number(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawRightString(
        doc.pagesize[0] - 0.8 * cm,
        0.55 * cm,
        f"Pagina {doc.page}",
    )
    canvas.restoreState()


def make_paragraph(text, style):
    return Paragraph(str(text), style)


def dataframe_chunks(df, rows_per_page=27):
    for start in range(0, len(df), rows_per_page):
        yield df.iloc[start : start + rows_per_page]


def build_excel_download(
    visible_order_df,
    total_units,
    total_rows,
    total_skus,
    generated_at,
    data_updated_at,
):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table as ExcelTable
    from openpyxl.worksheet.table import TableStyleInfo
    from openpyxl.utils import get_column_letter

    buffer = BytesIO()
    sheet_name = "Orden produccion"
    startrow = 4

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        visible_order_df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name,
            startrow=startrow,
        )

        worksheet = writer.sheets[sheet_name]

        max_row = worksheet.max_row
        max_col = worksheet.max_column
        last_col_letter = get_column_letter(max_col)
        header_row = startrow + 1
        first_data_row = header_row + 1

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        worksheet["A1"] = f"Orden de produccion - {TARGET_PERIOD_LABEL}"
        worksheet["A1"].font = Font(bold=True, size=14, color="1F2933")
        worksheet["A1"].alignment = Alignment(horizontal="left")

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        worksheet["A2"] = (
            f"Descarga/generacion: {format_bogota(generated_at)} Bogota | "
            f"Datos actualizados: {format_bogota(data_updated_at)} Bogota | "
            f"Periodo objetivo: {TARGET_PERIOD_LABEL} | "
            f"Unidades: {total_units:,} | "
            f"Filas: {total_rows:,} | "
            f"Referencias/tallas: {total_skus:,}"
        )
        worksheet["A2"].font = Font(size=9, color="4B5563")

        header_fill = PatternFill("solid", fgColor="243B53")
        header_font = Font(bold=True, color="FFFFFF")
        red_fill = PatternFill("solid", fgColor="F8D7DA")
        red_font = Font(bold=True, color="8A1F2D")
        yellow_fill = PatternFill("solid", fgColor="FFF3CD")
        yellow_font = Font(bold=True, color="664D03")
        thin_border = Border(
            left=Side(style="thin", color="CBD2D9"),
            right=Side(style="thin", color="CBD2D9"),
            top=Side(style="thin", color="CBD2D9"),
            bottom=Side(style="thin", color="CBD2D9"),
        )

        for cell in worksheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in worksheet.iter_rows(
            min_row=first_data_row,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
        ):
            is_total_row = str(row[0].value or "").strip().upper() == "TOTAL POR TALLA"
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="left" if cell.column == 1 else "right",
                    vertical="center",
                    wrap_text=cell.column == 1,
                )

                if is_total_row:
                    cell.fill = header_fill
                    cell.font = header_font
                    if cell.column > 1:
                        cell.number_format = '#,##0'
                elif cell.column > 1:
                    cell.number_format = '#,##0'
                    value = cell.value or 0
                    if value > 15:
                        cell.fill = red_fill
                        cell.font = red_font
                    elif value > 10:
                        cell.fill = yellow_fill
                        cell.font = yellow_font

        worksheet.column_dimensions["A"].width = 44
        for column_idx in range(2, max_col + 1):
            worksheet.column_dimensions[get_column_letter(column_idx)].width = 10

        table_ref = f"A{header_row}:{last_col_letter}{max_row}"
        excel_table = ExcelTable(displayName="OrdenProduccion", ref=table_ref)
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(excel_table)
        worksheet.freeze_panes = f"B{first_data_row}"
        worksheet.auto_filter.ref = table_ref

    buffer.seek(0)
    return buffer.getvalue()


def build_pdf_download(
    visible_order_df,
    total_units,
    total_rows,
    total_skus,
    generated_at,
    data_updated_at,
):
    buffer = BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.7 * cm,
        rightMargin=0.7 * cm,
        topMargin=0.7 * cm,
        bottomMargin=0.85 * cm,
        title=f"Orden de produccion {TARGET_PERIOD_LABEL}",
    )
    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height,
        id="normal",
    )
    doc.addPageTemplates(
        [
            PageTemplate(
                id="page",
                frames=[frame],
                onPage=pdf_page_number,
            )
        ]
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleDYUNIC",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        textColor=colors.HexColor("#1f2933"),
        alignment=TA_LEFT,
    )
    body_style = ParagraphStyle(
        "BodyDYUNIC",
        parent=styles["BodyText"],
        fontSize=7.5,
        leading=9,
    )
    header_style = ParagraphStyle(
        "HeaderDYUNIC",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=6.5,
        leading=8,
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    left_cell_style = ParagraphStyle(
        "LeftCellDYUNIC",
        parent=styles["BodyText"],
        fontSize=6.3,
        leading=7.5,
        alignment=TA_LEFT,
    )
    number_cell_style = ParagraphStyle(
        "NumberCellDYUNIC",
        parent=styles["BodyText"],
        fontSize=6.3,
        leading=7.5,
        alignment=TA_RIGHT,
    )

    story = []
    columns = list(visible_order_df.columns)

    for page_index, chunk in enumerate(dataframe_chunks(visible_order_df), start=1):
        title = f"Orden de produccion - {TARGET_PERIOD_LABEL}"
        if page_index > 1:
            title += " (continuacion)"

        story.append(make_paragraph(title, title_style))
        story.append(
            make_paragraph(
                (
                    f"Descarga/generacion: {format_bogota(generated_at)} Bogota | "
                    f"Datos actualizados: {format_bogota(data_updated_at)} Bogota | "
                    f"Periodo objetivo: {TARGET_PERIOD_LABEL} | "
                    f"Unidades: {total_units:,} | "
                    f"Filas: {total_rows:,} | "
                    f"Referencias/tallas: {total_skus:,}"
                ),
                body_style,
            )
        )
        story.append(Spacer(1, 0.12 * cm))

        data = [[make_paragraph(column.replace("_", " "), header_style) for column in columns]]
        for _, row in chunk.iterrows():
            data.append(
                [
                    make_paragraph(row[columns[0]], left_cell_style),
                    *[
                        make_paragraph(
                            "" if int(row[column]) == 0 else f"{int(row[column]):,}",
                            number_cell_style,
                        )
                        for column in columns[1:]
                    ],
                ]
            )

        first_width = 6.9 * cm
        remaining_width = doc.width - first_width
        size_width = max(0.75 * cm, remaining_width / max(1, len(columns) - 1))
        table = Table(
            data,
            repeatRows=1,
            colWidths=[first_width] + [size_width] * (len(columns) - 1),
        )
        commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#243b53")),
            ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#cbd2d9")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9fb")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ]

        for row_index, (_, row) in enumerate(chunk.iterrows(), start=1):
            if str(row[columns[0]]).strip().upper() == "TOTAL POR TALLA":
                commands.extend(
                    [
                        ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#243b53")),
                        ("TEXTCOLOR", (0, row_index), (-1, row_index), colors.white),
                        ("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold"),
                        ("LINEABOVE", (0, row_index), (-1, row_index), 0.8, colors.HexColor("#111827")),
                    ]
                )
                continue

            for column_index, column in enumerate(columns[1:], start=1):
                value = int(row[column])
                if value > 15:
                    commands.append(
                        (
                            "BACKGROUND",
                            (column_index, row_index),
                            (column_index, row_index),
                            colors.HexColor("#f8d7da"),
                        )
                    )
                    commands.append(
                        (
                            "TEXTCOLOR",
                            (column_index, row_index),
                            (column_index, row_index),
                            colors.HexColor("#8a1f2d"),
                        )
                    )
                elif value > 10:
                    commands.append(
                        (
                            "BACKGROUND",
                            (column_index, row_index),
                            (column_index, row_index),
                            colors.HexColor("#fff3cd"),
                        )
                    )
                    commands.append(
                        (
                            "TEXTCOLOR",
                            (column_index, row_index),
                            (column_index, row_index),
                            colors.HexColor("#664d03"),
                        )
                    )

        table.setStyle(TableStyle(commands))
        story.append(table)

        if page_index < math.ceil(len(visible_order_df) / 27):
            story.append(PageBreak())

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


if st.button("🔄 Refresh", type="primary"):
    st.rerun()

detail_df, order_df = build_order()

ventas_updated_at = file_updated_at_bogota(VENTAS_PATH)
inventario_updated_at = file_updated_at_bogota(INVENTARIO_PATH)
data_updated_at = max(ventas_updated_at, inventario_updated_at)

if detail_df.empty:
    st.warning("No hay necesidades de produccion con los datos actuales.")
    st.stop()

family_options = sorted(
    detail_df["familia_produccion"].dropna().astype(str).unique(),
    key=normalize_text,
)
selected_families = st.multiselect(
    "Familia de produccion",
    options=family_options,
    default=family_options,
    help="Este filtro usa la columna Produccion de TB INVENTARIO, relacionada por ID_BUSQUEDA.",
)

if selected_families:
    detail_df = detail_df[detail_df["familia_produccion"].isin(selected_families)].copy()
else:
    detail_df = detail_df.iloc[0:0].copy()

order_df = build_pivot(detail_df)

if detail_df.empty or order_df.empty:
    st.warning("No hay necesidades de produccion para las familias seleccionadas.")
    st.stop()

total_units = int(detail_df["orden_produccion"].sum())
total_rows = int(len(order_df))
total_skus = int(len(detail_df))

col1, col2, col3 = st.columns(3)

with col1:
    st.metric("Unidades proyectadas", f"{total_units:,}")

with col2:
    st.metric("Filas colegio/prenda", f"{total_rows:,}")

with col3:
    st.metric("Referencias/tallas", f"{total_skus:,}")

st.caption(
    f"Orden proyectada para {TARGET_PERIOD_LABEL} con inventario descontado, "
    "inventarios negativos tratados como cero y redondeo hacia arriba a multiplos de 5."
)
st.caption(
    f"Ultima actualizacion de datos: {format_bogota(data_updated_at)} Bogota | "
    f"ventas_df: {format_bogota(ventas_updated_at)} | "
    f"inventario: {format_bogota(inventario_updated_at)}"
)

numeric_columns = [
    column
    for column in order_df.columns
    if column != "colegio y prenda"
]

display_df = order_df.copy()
display_df[numeric_columns] = display_df[numeric_columns].astype(int)
display_df_with_total = add_total_row(display_df, "colegio y prenda")

pdf_generated_at = now_bogota()
pdf_bytes = build_pdf_download(
    display_df_with_total,
    total_units=total_units,
    total_rows=total_rows,
    total_skus=total_skus,
    generated_at=pdf_generated_at,
    data_updated_at=data_updated_at,
)
excel_bytes = build_excel_download(
    display_df_with_total,
    total_units=total_units,
    total_rows=total_rows,
    total_skus=total_skus,
    generated_at=pdf_generated_at,
    data_updated_at=data_updated_at,
)

download_col1, download_col2 = st.columns(2)

with download_col1:
    st.download_button(
        "📄 Descargar esta tabla en PDF carta horizontal",
        data=pdf_bytes,
        file_name=f"orden_produccion_{pdf_generated_at:%Y%m%d_%H%M%S}.pdf",
        mime="application/pdf",
        help=(
            "El PDF usa hora de Bogota, muestra la hora de generacion/descarga "
            "y la ultima actualizacion de datos."
        ),
        on_click="rerun",
    )

with download_col2:
    st.download_button(
        "📊 Descargar esta tabla en Excel",
        data=excel_bytes,
        file_name=f"orden_produccion_{pdf_generated_at:%Y%m%d_%H%M%S}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help=(
            "El Excel contiene la misma tabla visible, con filtros, panel congelado "
            "y colores de alerta."
        ),
        on_click="rerun",
    )

styled_df = (
    display_df_with_total
    .style
    .map(color_order_cells, subset=numeric_columns)
    .apply(style_total_row, axis=1)
    .format({column: "{:,.0f}" for column in numeric_columns})
)

st.dataframe(
    styled_df,
    width="stretch",
    height=720,
    hide_index=True,
)

with st.expander("Ver detalle de calculo"):
    detail_view = detail_df[
        [
            "colegio",
            "articulo",
            "familia_produccion",
            "talla",
            "inventario",
            "pronostico_cierre",
            "stock_minimo",
            "orden_sin_redondeo",
            "orden_produccion",
        ]
    ].copy()
    st.dataframe(
        detail_view,
        width="stretch",
        height=420,
        hide_index=True,
    )
